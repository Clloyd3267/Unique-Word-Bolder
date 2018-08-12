###################################################################################################
# Name        : UniqueWordBolder.py
# Author(s)   : Chris Lloyd, Andrew Southwick
# Description : A program to bold C&MA Bible Quizzing questions
# Github Link : https://github.com/Clloyd3267/Unique-Word-Bolder/
###################################################################################################

# External Imports
from pathlib import Path # Used for file manipulation
import openpyxl # For reading in verses and questions
import re # Used for pattern matching
import xlsxwriter # Used to write quizzes to excel files


class UniqueWordBolder:
    """
    A class to bold unique words.
    """

    def __init__(self, colLetters, questionsFileName = "Questions.xlsx", outputFilename = "Questions.xlsx"):
        """
        The constructor for class UniqueWordBolder.

        Parameters:
            colLetters (array of str): The columns letters to be bolded.
            questionsFileName (str): The output filename, defaults to "Questions.xlsx".
            outputFilename (str): The output filename, defaults to "Questions.xlsx".
        """

        self.debug = "Off" # A debug variable to enable / disable debug outputs
        self.allVerses = []  # An array to store all of the verses
        self.uniqueWords = []  # An array to hold all of the unique words
        self.cellsToBeBolded = [] # An array of the questions that were inputted


        self.importVerses() # Import the material to find unique words
        self.createUniqueWords() # Create Unique Words
        self.importQuestions(questionsFileName) # Import questions to be bolded
        self.exportQuestions(colLetters, outputFilename) # Export and bold questions

    ####################################################################################################################
    # Main Funcs
    ####################################################################################################################
    def importVerses(self, versesFileName = "Verses.xlsx"):
        """
        Function to import verses from excel file.

        Parameters:
            versesFileName (str): The input filename for verse list  (Defaults to "Verses.xlsx").

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code (All Verses): "A" or "a" or "On"
        """

        # Create the path for Verse file
        dataFilePath = Path("../Data Files/")  # Path where datafiles are stored

        if versesFileName == "Verses.xlsx":
            versesFilePath = dataFilePath / versesFileName
        else:
            versesFilePath = versesFileName

        # Try opening the verses file
        try:
            book = openpyxl.load_workbook(versesFilePath)
        except IOError:
            return "Error => Verses file does not exist!!!"

        sheet = book.worksheets[0]  # Open the first sheet

        # Read in and parse all verses
        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 4):
            # Check to make sure verse is valid
            verse = []
            valid = False

            for cell in row:
                if not cell.value:
                    verse.append("")
                else:
                    verse.append(str(cell.value).strip())
                    valid = True

            if not valid:
                continue
            if not verse[0]:
                return "Error => No Book!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]
            if not verse[1]:
                return "Error => No Chapter!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]
            if not verse[2]:
                return "Error => No Verse Number!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]
            if not verse[3]:
                return "Error => No Verse!!! " + verse[0] + " " + verse[1] + ":" + verse[2] + " " + verse[3]

            # Split verse
            verse.append(self.splitVerse(verse[3]))

            # Add verse to list of all verses
            self.allVerses.append(verse)

        # Print All Verses if debug enabled
        if self.debug != "Off" and ("A" in self.debug or "a" in self.debug or self.debug == "On"):
            print("")
            print("=== All Verses (" + str(len(self.allVerses)) + ") ===")
            for verse in self.allVerses:
                print(verse[0] + " " + verse[1] + ":" + verse[2] + " - " + verse[3])

        book.close()
        return 0 # Return with no errors

    def importQuestions(self, questionsFileName = "Questions.xlsx"):
        """
        Function to import questions to be bolded from an excel file.

        Parameters:
            questionsFileName (str): The input filename for questions  (Defaults to "Questions.xlsx").

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code (All Verses): "Q" or "q" or "On"
        """

        # Create the path for Questions file
        dataFilePath = Path("../Data Files/") # Path where datafiles are stored

        if questionsFileName == "Questions.xlsx":
            questionsFilePath = dataFilePath / questionsFileName
        else:
            questionsFilePath = questionsFileName

        # Try opening the Questions file
        try:
            book = openpyxl.load_workbook(questionsFilePath)
        except IOError:
            return "Error => Questions file does not exist!!!"

        sheet = book.active  # Open the active sheet

        # Loop through the sheet and pull out certain col
        for row in sheet.iter_rows(min_row = 1, min_col = 1):
            fields = []
            for cell in row:
                if cell.value == None:
                    fields.append("")
                else:
                    fields.append(str(cell.value))
            self.cellsToBeBolded.append(fields)
        book.close()

        # Print Questions if debug enabled
        if self.debug != "Off" and ("Q" in self.debug or "q" in self.debug or self.debug == "On"):
            print("")
            print("=== Questions to be Bolded (" + str(len(self.cellsToBeBolded)) + ") ===")
            for row in self.cellsToBeBolded:
                for cell in row:
                    print(cell, end = " - ")
                print("")

    def createUniqueWords(self):
        """
        Function to create list of all Unique Words.

        Returns:
            (0): No errors, (Anything else): Errors.

        Debug Code (All Verses): "U" or "u" or "On"
        """

        tempConcordance = {}
        for verse in self.allVerses:
            for i, word in enumerate(verse[4]):
                newVerseText = verse[3][0:word[1]] + "◆" + verse[3][word[1] + len(word[0]):]
                word = str(word[0]).upper()

                if word in tempConcordance:
                    tempConcordance[word][1].append([verse[0], verse[1], verse[2], newVerseText])
                    tempConcordance[word][0] += 1
                else:
                    tempConcordance[word] = [1, [[verse[0], verse[1], verse[2], newVerseText]]]

        for word, value in sorted(tempConcordance.items()):
            firstOccurence = value[1][0][0:3]
            uniqueWord = True
            for occurence in value[1]:
                if occurence[0:3] != firstOccurence:
                    uniqueWord = False
            if uniqueWord:
                self.uniqueWords.append(word.lower())

        # Print Unique Words if debug enabled
        if self.debug != "Off" and ("U" in self.debug or "u" in self.debug or self.debug == "On"):
            print("")
            print("=== Unique Words (" + str(len(self.uniqueWords)) + ") ===")
            for word, value in sorted(self.uniqueWords.items()):
                print(word + " - " + value[0] + " " + value[1] + ":" + value[2])

        return 0  # Return with no errors

    def exportQuestions(self, colNames, outputFilename = "Questions.xlsx"):
        """
        Function to export and bold questions to an Excel file.

        Parameters:
            colNames (array of str): The columns letters to be bolded.
            outputFilename (str): The output filename, defaults to "Questions.xlsx".

        Returns:
            (0): No errors, (Anything else): Errors.
        """

        # Convert the column letters to numbers
        colNums = []
        for col in colNames:
            colNums.append(ord(col.lower()[0]) - 97)

        # Create the output file
        if outputFilename == "Questions.xlsx":
            fileName = Path("../Questions.xlsx")
            workbook = xlsxwriter.Workbook(fileName)
        else:
            workbook = xlsxwriter.Workbook(outputFilename)

        # Set cell formats
        bold = workbook.add_format({'bold': 1})
        borderFormat = workbook.add_format({'border': 1})

        worksheet = workbook.add_worksheet("Questions")

        for rowNum, row in enumerate(self.cellsToBeBolded):
            for colNum, cell in enumerate(row):
                if colNum in colNums:
                    if cell != "":
                        worksheet.set_column(colNum, colNum, 50)
                        worksheet.write_rich_string(rowNum, colNum, *self.boldUniqueWords(cell, bold), borderFormat)
                    else:
                        worksheet.write_string(rowNum, colNum, cell, borderFormat)
                else:
                    worksheet.write_string(rowNum, colNum, cell, borderFormat)
        workbook.close()  # Close workbook
        return 0  # Return with no errors


    ####################################################################################################################
    # Helper Funcs
    ####################################################################################################################
    def splitVerse(self, verseText):
        """
        Function to split a verse into individual words.

        Parameters:
            verseText(str): Text of verse to be split.

        Returns:
            splitVerse(arr): Array of array of str and int representing split verse text.
        """
        splitVerse = []
        partOfWord = ["", 0]

        # Loop through all characters with index
        for i, character in enumerate(verseText):
            # Character is part of word
            if (character.isalnum()) or \
            (character == "-") or \
            (character in ["’" , "'"] and partOfWord[0] != "" and
            (i != 0 and (verseText[i - 5:i].lower() == "jesus") or
            (i != len(verseText) - 1 and i != 0 and verseText[i - 1].isalnum() and verseText[i + 1].isalnum()))):
                if partOfWord[0] == "":
                    partOfWord[1] = i
                partOfWord[0] += character

            # Character is not part of word
            else:
                if partOfWord[0]:
                    splitVerse.append(partOfWord.copy())
                    partOfWord[0] = ""
                    partOfWord[1] = 0

        # Append any leftover characters to splitVerse
        if partOfWord[0]:
            splitVerse.append(partOfWord.copy())

        return splitVerse

    def boldUniqueWords(self, myString, boldFormat):
        """
        Function to bold unique words in a particular string.

        Parameters:
            myString (str): The input string to be bolded.
            boldFormat (xlsxwriter format object): The format to be applied to unique words.

        Returns:
            result (array) Array of strings and xlsxwriter objects.
        """

        # Check to make sure string is not a reference question or a quote
        rMatch = re.search(r'According\sto.*Chapter', myString, re.IGNORECASE)
        qMatch = re.search(r'Quote\sto.*Chapter', myString, re.IGNORECASE)
        if rMatch or qMatch:
            return [myString]

        result = []
        start = 0
        splitVerse = self.splitVerse(myString)
        for word in splitVerse:
            if word[0].lower() in self.uniqueWords:
                result.append(myString[start:word[1]])
                result.append(boldFormat)
                result.append(myString[word[1]:word[1] + len(word[0])])
                start = len(word[0]) + word[1]
        if start != len(myString):
            result.append(myString[start:])
        return result


if __name__ == "__main__":
    uLB = UniqueWordBolder(["F", "G"])